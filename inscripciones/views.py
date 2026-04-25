import csv
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from .models import FormularioInscripcionHUT, Pais, ProvinciaEstado, CursoHUT, GruposMoodle
from .forms import InscripcionForm, GruposMoodleForm, CursoHUTForm, InscriptoLoginForm, SeleccionGrupoForm, VoluntarioForm
from .envio_mail import enviar_confirmacion_inscripcion


@login_required
def inscripcion_lista(request):
    """Vista principal: grilla de inscripciones (solo administradores logueados)."""
    query = request.GET.get('q', '').strip()
    curso_id = request.GET.get('curso')
    pais_id = request.GET.get('pais')
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    inscripciones = FormularioInscripcionHUT.objects.select_related('pais', 'provincia', 'curso').all()

    if query:
        inscripciones = inscripciones.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(email__icontains=query) |
            Q(congregacion__icontains=query) |
            Q(ciudad__icontains=query) |
            Q(pastor__icontains=query)
        )

    if curso_id:
        inscripciones = inscripciones.filter(curso_id=curso_id)
        
    if pais_id:
        inscripciones = inscripciones.filter(pais_id=pais_id)
        
    if fecha_desde_str:
        fecha_desde = f"{fecha_desde_str} 00:00:00"
        inscripciones = inscripciones.filter(fecha_creacion__gte=fecha_desde)
        
    if fecha_hasta_str:
        fecha_hasta = f"{fecha_hasta_str} 23:59:59"
        inscripciones = inscripciones.filter(fecha_creacion__lte=fecha_hasta)

    inscripciones = inscripciones.order_by('-fecha_creacion')

    paginator = Paginator(inscripciones, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
        'curso_id_filtro': curso_id,
        'pais_id_filtro': pais_id,
        'fecha_desde': fecha_desde_str,
        'fecha_hasta': fecha_hasta_str,
        'total': inscripciones.count(),
        'cursos': CursoHUT.objects.all().order_by('-anio', '-id'),
        'paises': Pais.objects.all().order_by('paisNombre'),
    }
    return render(request, 'inscripciones/lista.html', context)


def inscripcion_crear(request):
    """Formulario público para crear una nueva inscripción (sin login requerido)."""
    curso_activo = CursoHUT.objects.filter(activo=True).first()
    if curso_activo and not curso_activo.inscripciones_abiertas:
        return inscripciones_cerradas(request)

    if request.method == 'POST':
        form = InscripcionForm(request.POST)
        if form.is_valid():
            inscripcion = form.save()
            enviar_confirmacion_inscripcion(inscripcion)
            # Guardar el ID en sesión para mostrar datos del grupo en la página de agradecimiento
            request.session['inscripcion_exitosa_id'] = inscripcion.id
            return redirect('inscripcion_gracias')
    else:
        form = InscripcionForm()

    return render(request, 'publico/formulario_publico.html', {'form': form})


def inscripcion_gracias(request):
    """Página de agradecimiento post-inscripción (pública)."""
    inscripcion_id = request.session.pop('inscripcion_exitosa_id', None)
    context = {}
    if inscripcion_id:
        try:
            inscripcion = FormularioInscripcionHUT.objects.select_related('grupo').get(id=inscripcion_id)
            context['inscripcion'] = inscripcion
            if inscripcion.grupo:
                context['grupo'] = inscripcion.grupo
                context['whatsapp_url'] = inscripcion.grupo.url_whatsapp
        except FormularioInscripcionHUT.DoesNotExist:
            pass
    return render(request, 'publico/gracias_publica.html', context)


def inscripciones_cerradas(request):
    """Página que se muestra cuando las inscripciones están cerradas (pública)."""
    return render(request, 'publico/inscripciones_cerradas.html')


def get_provincias(request, pais_id):
    """API JSON: devuelve provincias filtradas por país (para AJAX)."""
    provincias = ProvinciaEstado.objects.filter(idPais_id=pais_id).order_by('provinciaNombre')
    data = [{'id': p.id, 'nombre': p.provinciaNombre} for p in provincias]
    return JsonResponse(data, safe=False)



@login_required
def exportar_csv_moodle(request):
    """Exporta inscripciones del curso activo como XLSX compatible con Moodle."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    curso_activo = CursoHUT.objects.filter(activo=True).first()
    if not curso_activo:
        return HttpResponse("No hay curso activo.")

    inscripciones = (
        FormularioInscripcionHUT.objects
        .select_related('grupo')
        .filter(curso=curso_activo)
        .order_by('apellido', 'nombre')
    )

    curso_nombre = curso_activo.nombre.replace(' ', '')
    fecha = datetime.now().strftime('%Y%m%d')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Moodle'

    # Estilos para el header
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3D6B67', end_color='3D6B67', fill_type='solid')
    header_align = Alignment(horizontal='center')

    headers = ['username', 'firstname', 'lastname', 'email', 'course1', 'role1', 'group1']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Datos
    for row, i in enumerate(inscripciones, 2):
        grupo_nombre = i.grupo.nombre if i.grupo else ""
        ws.cell(row=row, column=1, value=i.email)
        ws.cell(row=row, column=2, value=i.nombre)
        ws.cell(row=row, column=3, value=i.apellido)
        ws.cell(row=row, column=4, value=i.email)
        ws.cell(row=row, column=5, value=curso_nombre)
        ws.cell(row=row, column=6, value='student')
        ws.cell(row=row, column=7, value=grupo_nombre)

    # Auto-ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f'MOODLE_{curso_nombre}_{fecha}.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response

@login_required
def exportar_csv_completo(request):
    """Exporta TODAS las inscripciones (filtradas o no) a un archivo XLSX con todos los datos."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Aplicamos la misma lógica de búsqueda que en la lista
    query = request.GET.get('q', '').strip()
    curso_id = request.GET.get('curso')
    pais_id = request.GET.get('pais')
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    inscripciones = FormularioInscripcionHUT.objects.select_related('pais', 'provincia', 'curso', 'grupo').all()

    if query:
        inscripciones = inscripciones.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(email__icontains=query) |
            Q(congregacion__icontains=query) |
            Q(ciudad__icontains=query) |
            Q(pastor__icontains=query)
        )
        
    if curso_id:
        inscripciones = inscripciones.filter(curso_id=curso_id)
        
    if pais_id:
        inscripciones = inscripciones.filter(pais_id=pais_id)
        
    if fecha_desde_str:
        fecha_desde = f"{fecha_desde_str} 00:00:00"
        inscripciones = inscripciones.filter(fecha_creacion__gte=fecha_desde)
        
    if fecha_hasta_str:
        fecha_hasta = f"{fecha_hasta_str} 23:59:59"
        inscripciones = inscripciones.filter(fecha_creacion__lte=fecha_hasta)

    inscripciones = inscripciones.order_by('-fecha_creacion')
        
    fecha = datetime.now().strftime('%Y%m%d_%H%M')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inscripciones Completas'

    # Estilos para el header
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4744', end_color='1F4744', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = [
        'ID', 'Fecha Inscripción', 'Curso', 'Nombre', 'Apellido', 'Email', 'Teléfono', 
        'Edad', 'Estado Civil', 'País', 'Provincia', 'Ciudad', 'Congregación', 
        'Pastor', 'Estado', 'Código Acceso', 'Grupo Asignado'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Datos
    for row, i in enumerate(inscripciones, 2):
        # Determinar estado
        estado = "Activo"
        if i.abandonado:
            estado = "Abandonado"
        elif i.fecha_finalizacion:
            estado = "Finalizado"

        grupo_nombre = i.grupo.nombre if i.grupo else "Sin asignar"
        curso_nombre = i.curso.nombre if i.curso else ""
        fecha_str = i.fecha_creacion.strftime('%d/%m/%Y %H:%M') if i.fecha_creacion else ""

        ws.cell(row=row, column=1, value=i.id)
        ws.cell(row=row, column=2, value=fecha_str)
        ws.cell(row=row, column=3, value=curso_nombre)
        ws.cell(row=row, column=4, value=i.nombre)
        ws.cell(row=row, column=5, value=i.apellido)
        ws.cell(row=row, column=6, value=i.email)
        ws.cell(row=row, column=7, value=i.telefono)
        ws.cell(row=row, column=8, value=i.edad)
        ws.cell(row=row, column=9, value=i.estadoCivil)
        ws.cell(row=row, column=10, value=i.pais.paisNombre if i.pais else "")
        ws.cell(row=row, column=11, value=i.provincia.provinciaNombre if i.provincia else "")
        ws.cell(row=row, column=12, value=i.ciudad)
        ws.cell(row=row, column=13, value=i.congregacion)
        ws.cell(row=row, column=14, value=i.pastor)
        ws.cell(row=row, column=15, value=estado)
        ws.cell(row=row, column=16, value=i.codigo_acceso)
        ws.cell(row=row, column=17, value=grupo_nombre)

    # Auto-ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40) # Limitar ancho maximo a 40 para que no sea inmenso

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f'InscripcionesHUT_Completas_{fecha}.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ==============================================================================
# CURSOS HUT CRUD
# ==============================================================================

class CursoHUTListView(LoginRequiredMixin, ListView):
    model = CursoHUT
    template_name = 'inscripciones/cursohut_list.html'
    context_object_name = 'cursos'
    paginate_by = 15

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) |
                Q(anio__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        return context

class CursoHUTCreateView(LoginRequiredMixin, CreateView):
    model = CursoHUT
    form_class = CursoHUTForm
    template_name = 'inscripciones/cursohut_form.html'
    success_url = reverse_lazy('cursos_lista')

    def form_valid(self, form):
        messages.success(self.request, 'Curso creado correctamente.')
        return super().form_valid(form)

class CursoHUTUpdateView(LoginRequiredMixin, UpdateView):
    model = CursoHUT
    form_class = CursoHUTForm
    template_name = 'inscripciones/cursohut_form.html'
    success_url = reverse_lazy('cursos_lista')

    def form_valid(self, form):
        messages.success(self.request, 'Curso actualizado correctamente.')
        return super().form_valid(form)

class CursoHUTDeleteView(LoginRequiredMixin, DeleteView):
    model = CursoHUT
    template_name = 'inscripciones/cursohut_confirm_delete.html'
    success_url = reverse_lazy('cursos_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Curso eliminado correctamente.')
        return super().delete(request, *args, **kwargs)

@login_required
def toggle_inscripciones_curso(request, pk):
    from django.shortcuts import get_object_or_404
    curso = get_object_or_404(CursoHUT, pk=pk)
    if curso.activo:
        curso.inscripciones_abiertas = not curso.inscripciones_abiertas
        curso.save(update_fields=['inscripciones_abiertas'])
        estado = "abiertas" if curso.inscripciones_abiertas else "cerradas"
        messages.success(request, f'Inscripciones {estado} para el curso {curso.nombre}.')
    else:
        messages.error(request, 'Solo se pueden abrir o cerrar inscripciones en el curso activo.')
    
    return redirect('cursos_lista')

# ==============================================================================
# GRUPOS MOODLE CRUD
# ==============================================================================

class GruposMoodleListView(LoginRequiredMixin, ListView):
    model = GruposMoodle
    template_name = 'inscripciones/gruposmoodle_list.html'
    context_object_name = 'grupos'
    paginate_by = 15

    def get_queryset(self):
        queryset = super().get_queryset().select_related('curso')
        q = self.request.GET.get('q')
        curso_id = self.request.GET.get('curso')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) |
                Q(tutor__icontains=q) |
                Q(dia__icontains=q)
            )
        if curso_id:
            queryset = queryset.filter(curso_id=curso_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        context['curso_id_filtro'] = self.request.GET.get('curso', '')
        context['cursos'] = CursoHUT.objects.all().order_by('-anio', '-id')
        for grupo in context['grupos']:
            # Usar el campo participantes directamente del modelo GrupoMoodle,
            # el cual ya se mantiene sincronizado mediante las señales pre_save y post_save.
            grupo.cantidad_inscriptos = grupo.participantes
        return context

class GruposMoodleCreateView(LoginRequiredMixin, CreateView):
    model = GruposMoodle
    form_class = GruposMoodleForm
    template_name = 'inscripciones/gruposmoodle_form.html'
    success_url = reverse_lazy('grupos_lista')

    def form_valid(self, form):
        messages.success(self.request, 'Grupo creado correctamente.')
        return super().form_valid(form)

class GruposMoodleUpdateView(LoginRequiredMixin, UpdateView):
    model = GruposMoodle
    form_class = GruposMoodleForm
    template_name = 'inscripciones/gruposmoodle_form.html'
    success_url = reverse_lazy('grupos_lista')

    def form_valid(self, form):
        messages.success(self.request, 'Grupo actualizado correctamente.')
        return super().form_valid(form)

class GruposMoodleDeleteView(LoginRequiredMixin, DeleteView):
    model = GruposMoodle
    template_name = 'inscripciones/gruposmoodle_confirm_delete.html'
    success_url = reverse_lazy('grupos_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Grupo eliminado correctamente.')
        return super().delete(request, *args, **kwargs)

# ==============================================================================
# FLUJO DE SELECCIÓN DE GRUPO (PÚBLICO)
# ==============================================================================

def inscripto_login_view(request):
    """Vista pública para que el inscripto ingrese su Código Único y Email."""
    if request.method == 'POST':
        form = InscriptoLoginForm(request.POST)
        if form.is_valid():
            codigo_acceso = form.cleaned_data['codigo_acceso'].strip()
            email = form.cleaned_data['email'].strip()

            try:
                inscripcion = FormularioInscripcionHUT.objects.get(codigo_acceso__iexact=codigo_acceso, email__iexact=email)
                request.session['inscripto_id'] = inscripcion.id
                return redirect('inscripto_seleccionar_grupo')
            except FormularioInscripcionHUT.DoesNotExist:
                messages.error(request, 'No se encontró ninguna inscripción con esa combinación de Código y Correo Electrónico.')
    else:
        form = InscriptoLoginForm()
    
    return render(request, 'publico/login_inscripto.html', {'form': form})

def inscripto_seleccion_grupo_view(request):
    """Vista pública donde el inscripto elige su grupo de la lista disponible."""
    inscripto_id = request.session.get('inscripto_id')
    if not inscripto_id:
        return redirect('inscripto_login')

    try:
        inscripcion = FormularioInscripcionHUT.objects.get(id=inscripto_id)
    except FormularioInscripcionHUT.DoesNotExist:
        del request.session['inscripto_id']
        return redirect('inscripto_login')

    if request.method == 'POST':
        form = SeleccionGrupoForm(request.POST, instance=inscripcion)
        if form.is_valid():
            form.save()
            
            # Enviar email con el link del grupo
            from .envio_mail import enviar_confirmacion_grupo
            enviar_confirmacion_grupo(inscripcion, inscripcion.grupo)
            
            # Limpiar sesión después de éxito
            if 'inscripto_id' in request.session:
                del request.session['inscripto_id']
            return redirect('inscripto_grupo_exito')
    else:
        form = SeleccionGrupoForm(instance=inscripcion)

    return render(request, 'publico/seleccionar_grupo.html', {'form': form, 'inscripcion': inscripcion})

def inscripto_grupo_exito_view(request):
    """Vista de éxito tras elegir grupo."""
    return render(request, 'publico/exito_grupo.html')

# ==============================================================================
# FORMULARIO VOLUNTARIOS (PÚBLICO)
# ==============================================================================

def formulario_voluntario_crear(request):
    """Formulario público para registrar un Voluntario."""
    from .models import TerminosCondiciones
    import json

    if request.method == 'POST':
        form = VoluntarioForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('formulario_voluntario_gracias')
    else:
        form = VoluntarioForm()

    # Cargar términos y condiciones desde la base de datos
    terminos_qs = TerminosCondiciones.objects.all()
    terminos_dict = {t.rol: t.texto_html for t in terminos_qs}
    terminos_json = json.dumps(terminos_dict)

    return render(request, 'publico/formulario_voluntario.html', {
        'form': form,
        'terminos_json': terminos_json,
    })

def formulario_voluntario_gracias(request):
    """Página de agradecimiento post-inscripción de voluntarios."""
    return render(request, 'publico/gracias_voluntario.html')

# ==============================================================================
# IMPORTACIÓN DE ALUMNOS (ADMIN)
# ==============================================================================

@login_required
def cambio_estado_alumno_view(request):
    """Vista para importar un CSV y actualizar estado de alumnos (fecha de finalización y abandonos)."""
    from .forms import ImportarEstadoAlumnosForm
    from django.utils import timezone
    from django.urls import reverse
    import csv

    inscripciones_curso = None
    curso_seleccionado = None

    if request.method == 'POST':
        form = ImportarEstadoAlumnosForm(request.POST, request.FILES)
        if form.is_valid():
            curso_seleccionado = form.cleaned_data['curso']
            archivo_csv = request.FILES['archivo_csv']
            
            try:
                decoded_file = archivo_csv.read().decode('utf-8-sig').splitlines()
                reader = csv.reader(decoded_file, delimiter=';')
                headers = next(reader)  # Ignorar la cabecera
            except Exception as e:
                messages.error(request, f"Error al leer el archivo CSV: {e}")
                return redirect('cambio_estado_alumno')

            datos_csv = {}
            for row in reader:
                if not row:
                    continue
                # Se asume formato: Nombre, Apellido, mail, curso, fecha de finalizacion
                if len(row) >= 5:
                    email_val = row[2].strip().lower()
                    fecha_val = row[4].strip()
                    datos_csv[email_val] = fecha_val

            inscripciones = FormularioInscripcionHUT.objects.filter(curso=curso_seleccionado)
            
            actualizados_finalizacion = 0
            marcados_abandonados = 0

            for inscripcion in inscripciones:
                email_db = inscripcion.email.lower().strip()
                if email_db in datos_csv:
                    fecha_str = datos_csv[email_db]
                    if fecha_str:
                        try:
                            if '-' in fecha_str:
                                dt = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                            elif '/' in fecha_str:
                                dt = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                            else:
                                dt = None

                            if dt:
                                inscripcion.fecha_finalizacion = dt
                                inscripcion.abandonado = False
                                inscripcion.fecha_abandono = None
                                inscripcion.save(update_fields=['fecha_finalizacion', 'abandonado', 'fecha_abandono'])
                                actualizados_finalizacion += 1
                        except ValueError:
                            messages.warning(request, f"Formato de fecha inválido para {email_db}: {fecha_str}")
                else:
                    if not inscripcion.abandonado:
                        inscripcion.abandonado = True
                        inscripcion.fecha_abandono = timezone.now()
                        inscripcion.save(update_fields=['abandonado', 'fecha_abandono'])
                        marcados_abandonados += 1
            
            messages.success(request, f"Importación exitosa. {actualizados_finalizacion} finalizados, {marcados_abandonados} marcados como abandonados.")
            return redirect(f"{reverse('cambio_estado_alumno')}?curso={curso_seleccionado.id}")
    else:
        curso_id = request.GET.get('curso')
        initial = {}
        if curso_id:
            try:
                curso_seleccionado = CursoHUT.objects.get(id=curso_id)
                initial['curso'] = curso_seleccionado
                inscripciones_curso = FormularioInscripcionHUT.objects.filter(curso=curso_seleccionado).order_by('apellido', 'nombre')
            except CursoHUT.DoesNotExist:
                pass
        form = ImportarEstadoAlumnosForm(initial=initial)
        
    context = {
        'form': form,
        'inscripciones': inscripciones_curso,
        'curso_seleccionado': curso_seleccionado,
    }
    return render(request, 'inscripciones/cambio_estado_alumno.html', context)